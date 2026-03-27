"""
Generates sequence-validation.xlsx — comprehensive Excel validation workbook for SEQUENCE().

All 82 test rows now have confirmed expected values from real Excel (desktop, Microsoft 365).
The workbook auto-validates every row — open in Excel and check Pass/Fail column.

Usage:
    pip install openpyxl
    python3 scripts/gen-sequence-xlsx.py

Spec reference:
    https://support.microsoft.com/en-us/office/sequence-function-57467a98-57e0-4817-9f14-2eb78519ca90

Syntax: =SEQUENCE(rows, [columns], [start], [step])
Defaults: columns=1, start=1, step=1
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUTPUT_PATH = 'test/hyperformula-tests/compatibility/test_data/sequence-validation.xlsx'

# openpyxl needs _xlfn. prefix for newer Excel functions
FN = '_xlfn.SEQUENCE'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'SEQUENCE Validation'

# ── Styles ───────────────────────────────────────────────────────────────────
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
pass_fill = PatternFill('solid', fgColor='C6EFCE')
pass_font = Font(color='006100')
fail_fill = PatternFill('solid', fgColor='FFC7CE')
fail_font = Font(color='9C0006')
info_fill = PatternFill('solid', fgColor='FCE4EC')  # pink — manual/info only
group_font = Font(bold=True, size=12)
note_font = Font(italic=True, color='666666')
thin_border = Border(
    bottom=Side(style='thin', color='D9D9D9'),
)

# ── Setup Area (columns J-K) ────────────────────────────────────────────────
ws['J1'] = 'SETUP AREA'
ws['J1'].font = Font(bold=True, size=14)
setup_data = [
    ('J3', 'K3', 'Number 3', 3),
    ('J4', 'K4', 'Number 5', 5),
    ('J5', 'K5', 'Number 0', 0),
    ('J6', 'K6', 'Number -1', -1),
    ('J7', 'K7', 'Number 0.5', 0.5),
    ('J8', 'K8', 'Number 2.7', 2.7),
    ('J9', 'K9', 'Boolean TRUE', True),
    ('J10', 'K10', 'Boolean FALSE', False),
    ('J11', 'K11', 'Text "3"', '3'),
    ('J12', 'K12', 'Text "abc"', 'abc'),
    ('J13', 'K13', 'Empty cell', None),
    ('J14', 'K14', 'Formula empty =""', '=""'),
    ('J15', 'K15', 'Error #N/A', '=NA()'),
    ('J16', 'K16', 'Error #DIV/0!', '=1/0'),
    ('J17', 'K17', 'Number 1.9', 1.9),
    ('J18', 'K18', 'Number -2.7', -2.7),
]
for label_cell, val_cell, label, value in setup_data:
    ws[label_cell] = label
    ws[label_cell].font = Font(italic=True)
    if isinstance(value, str) and value.startswith('='):
        ws[val_cell] = value
    else:
        ws[val_cell] = value

# ── Header row ───────────────────────────────────────────────────────────────
headers = ['#', 'Group', 'Test Description', 'Formula (text)', 'Expected', 'Actual (formula)', 'Pass/Fail', 'Notes']
col_widths = [5, 20, 40, 45, 20, 20, 10, 40]
for col, (text, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=2, column=col, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Test cases ───────────────────────────────────────────────────────────────
# Format: (group, description, formula, expected, error_type, notes)
#
# expected: numeric/string value for value tests, display string for error tests
# error_type: None for value tests, ERROR.TYPE code for error tests:
#   2=#DIV/0!, 3=#VALUE!, 6=#NUM!, 7=#N/A, 9=#SPILL!, 14=#CALC!
# For manual/info rows (no auto-validation): formula=None
#
# Pass/Fail logic:
#   Value tests:  =IF(F=E, "PASS", "FAIL")
#   Error tests:  =IF(ISERROR(F), IF(ERROR.TYPE(F)=error_type, "PASS", "FAIL"), "FAIL")

tests = [
    # ── GROUP 1: Core sanity ────────────────────────────────────────────────
    ('Core sanity', 'Basic 4 rows', f'={FN}(4)', 1, None, 'Top-left of 4x1 array [1,2,3,4]'),
    ('Core sanity', '4x5 grid top-left', f'={FN}(4,5)', 1, None, 'Top-left of 4x5 grid'),
    ('Core sanity', '4x5 grid last cell', f'=INDEX({FN}(4,5),4,5)', 20, None, 'Bottom-right = rows*cols'),
    ('Core sanity', 'Start=10', f'={FN}(3,1,10)', 10, None, 'First value is start'),
    ('Core sanity', 'Start=10, step=5', f'={FN}(3,1,10,5)', 10, None, 'Sequence: 10,15,20'),
    ('Core sanity', 'Start=10, step=5 last', f'=INDEX({FN}(3,1,10,5),3,1)', 20, None, '10+2*5=20'),
    ('Core sanity', 'Single cell 1x1', f'={FN}(1,1)', 1, None, '1x1 returns scalar 1'),
    ('Core sanity', 'Single cell with start', f'={FN}(1,1,42)', 42, None, '1x1 start=42 -> 42'),

    # ── GROUP 2: Default parameters ─────────────────────────────────────────
    ('Defaults', 'cols omitted -> 1 col', f'=ROWS({FN}(3))', 3, None, '3 rows'),
    ('Defaults', 'cols omitted -> 1 col width', f'=COLUMNS({FN}(3))', 1, None, '1 column'),
    ('Defaults', 'start omitted -> 1', f'={FN}(3,2)', 1, None, 'Default start=1'),
    ('Defaults', 'step omitted -> 1', f'={FN}(3,2,0)', 0, None, 'start=0, step defaults to 1 -> 0,1,2,...'),
    ('Defaults', 'step omitted last value', f'=INDEX({FN}(3,2,0),3,2)', 5, None, '0 + (3*2-1)*1 = 5'),

    # ── GROUP 3: Empty args (emptyAsDefault) ────────────────────────────────
    ('Empty args', 'cols empty -> default 1', f'={FN}(3,)', 1, None, 'Empty cols -> 1'),
    ('Empty args', 'cols empty rows check', f'=ROWS({FN}(3,))', 3, None, '3 rows'),
    ('Empty args', 'cols empty cols check', f'=COLUMNS({FN}(3,))', 1, None, '1 column'),
    ('Empty args', 'start empty -> default 1', f'={FN}(3,2,)', 1, None, 'Empty start -> 1'),
    ('Empty args', 'step empty -> default 1', f'={FN}(3,2,1,)', 1, None, 'Empty step -> 1'),
    ('Empty args', 'step empty last value', f'=INDEX({FN}(3,2,1,),3,2)', 6, None, '1 + 5*1 = 6'),
    ('Empty args', 'all optional empty', f'={FN}(3,,,)', 1, None, 'All defaults: cols=1, start=1, step=1'),
    ('Empty args', 'all optional empty last', f'=INDEX({FN}(3,,,),3,1)', 3, None, '1 + 2*1 = 3'),

    # ── GROUP 4: Negative & zero step ───────────────────────────────────────
    ('Step variants', 'step=0 (constant)', f'={FN}(3,1,5,0)', 5, None, 'All values = 5'),
    ('Step variants', 'step=0 last', f'=INDEX({FN}(3,1,5,0),3,1)', 5, None, 'Constant 5'),
    ('Step variants', 'negative step', f'={FN}(3,1,10,-3)', 10, None, '10, 7, 4'),
    ('Step variants', 'negative step last', f'=INDEX({FN}(3,1,10,-3),3,1)', 4, None, '10 + 2*(-3) = 4'),
    ('Step variants', 'negative start', f'={FN}(3,1,-5,2)', -5, None, '-5, -3, -1'),
    ('Step variants', 'fractional step', f'={FN}(4,1,0,0.5)', 0, None, '0, 0.5, 1, 1.5'),
    ('Step variants', 'fractional step last', f'=INDEX({FN}(4,1,0,0.5),4,1)', 1.5, None, '0 + 3*0.5 = 1.5'),

    # ── GROUP 5: Truncation of rows/cols ────────────────────────────────────
    ('Truncation', 'rows=2.7 truncates to 2', f'=ROWS({FN}(2.7,1))', 2, None, 'trunc(2.7)=2'),
    ('Truncation', 'rows=2.9 truncates to 2', f'=ROWS({FN}(2.9))', 2, None, 'trunc(2.9)=2'),
    ('Truncation', 'cols=3.5 truncates to 3', f'=COLUMNS({FN}(1,3.5))', 3, None, 'trunc(3.5)=3'),
    ('Truncation', 'rows=1.1 -> 1', f'=ROWS({FN}(1.1))', 1, None, 'trunc(1.1)=1'),
    ('Truncation', 'negative frac rows=-2.7', f'={FN}(-2.7)', '#VALUE!', 3, 'trunc(-2.7)=-2, negative -> #VALUE!'),
    ('Truncation', 'rows=0.9 -> 0 -> #CALC!', f'={FN}(0.9)', '#CALC!', 14, 'trunc(0.9)=0, zero dim -> #CALC!'),
    ('Truncation', 'cols=0.5 -> 0 -> #CALC!', f'={FN}(1,0.5)', '#CALC!', 14, 'trunc(0.5)=0, zero dim -> #CALC!'),

    # ── GROUP 6: Error conditions ───────────────────────────────────────────
    ('Errors', 'rows=0 -> #CALC!', f'={FN}(0)', '#CALC!', 14, 'Zero dim -> #CALC!'),
    ('Errors', 'rows=-1 -> #VALUE!', f'={FN}(-1)', '#VALUE!', 3, 'Negative dim -> #VALUE!'),
    ('Errors', 'cols=0 -> #CALC!', f'={FN}(1,0)', '#CALC!', 14, 'Zero dim -> #CALC!'),
    ('Errors', 'cols=-1 -> #VALUE!', f'={FN}(1,-1)', '#VALUE!', 3, 'Negative dim -> #VALUE!'),
    ('Errors', 'rows=0, cols=0 -> #CALC!', f'={FN}(0,0)', '#CALC!', 14, 'Zero dim -> #CALC!'),
    ('Errors', 'no args (syntax error)', None, None, None, 'Excel rejects at parse time — not a valid formula'),
    ('Errors', 'too many args (syntax error)', None, None, None, 'Excel rejects at parse time — not a valid formula'),
    ('Errors', 'rows=text "abc" -> #VALUE!', f'={FN}("abc")', '#VALUE!', 3, 'Non-numeric text -> #VALUE!'),
    ('Errors', 'cols=text "abc" -> #VALUE!', f'={FN}(3,"abc")', '#VALUE!', 3, 'Non-numeric text -> #VALUE!'),
    ('Errors', 'start=text "abc" -> #VALUE!', f'={FN}(3,1,"abc")', '#VALUE!', 3, 'Non-numeric text -> #VALUE!'),
    ('Errors', 'step=text "abc" -> #VALUE!', f'={FN}(3,1,1,"abc")', '#VALUE!', 3, 'Non-numeric text -> #VALUE!'),
    ('Errors', 'rows=#N/A -> propagates', f'={FN}(NA())', '#N/A', 7, 'Error propagation: #N/A'),
    ('Errors', 'start=#DIV/0! -> propagates', f'={FN}(3,1,1/0)', '#DIV/0!', 2, 'Error propagation: #DIV/0!'),

    # ── GROUP 7: Type coercion ──────────────────────────────────────────────
    ('Type coercion', 'rows=TRUE -> 1', f'={FN}(TRUE)', 1, None, 'TRUE coerces to 1, single cell'),
    ('Type coercion', 'rows=FALSE -> 0 -> #CALC!', f'={FN}(FALSE)', '#CALC!', 14, 'FALSE->0, zero dim -> #CALC!'),
    ('Type coercion', 'cols=TRUE -> 1', f'=COLUMNS({FN}(3,TRUE))', 1, None, 'TRUE coerces to 1 column'),
    ('Type coercion', 'start=TRUE -> 1', f'={FN}(1,1,TRUE)', 1, None, 'TRUE coerces to 1'),
    ('Type coercion', 'step=TRUE -> 1', f'={FN}(3,1,1,TRUE)', 1, None, 'TRUE coerces to 1'),
    ('Type coercion', 'step=FALSE -> 0', f'={FN}(3,1,5,FALSE)', 5, None, 'FALSE->0, constant array all=5'),
    ('Type coercion', 'rows="3" (numeric string)', f'=ROWS({FN}("3"))', 3, None, '"3" coerces to 3 rows'),
    ('Type coercion', 'start="10" string', f'={FN}(1,1,"10")', 10, None, '"10" coerces to 10'),
    ('Type coercion', 'rows=cell ref (K3=3)', f'=ROWS({FN}(K3))', 3, None, 'Cell ref with number works'),
    ('Type coercion', 'rows=empty cell ref -> #CALC!', f'={FN}(K13)', '#CALC!', 14, 'Empty cell -> 0, zero dim -> #CALC!'),
    ('Type coercion', 'rows="" formula -> #VALUE!', f'={FN}(K14)', '#VALUE!', 3, '="" -> #VALUE!'),

    # ── GROUP 8: Large sequences ────────────────────────────────────────────
    ('Large sequences', '100x100 top-left', f'={FN}(100,100)', 1, None, '10000 element array'),
    ('Large sequences', '100x100 last', f'=INDEX({FN}(100,100),100,100)', 10000, None, '1 + 9999*1'),
    ('Large sequences', '1000x1 last', f'=INDEX({FN}(1000),1000,1)', 1000, None, '1000th element'),
    ('Large sequences', '1x1000 last', f'=INDEX({FN}(1,1000),1,1000)', 1000, None, '1000th element'),

    # ── GROUP 9: Fill order (row-major) ─────────────────────────────────────
    ('Fill order', '2x3 [1,1]=1', f'=INDEX({FN}(2,3),1,1)', 1, None, 'Row-major fill'),
    ('Fill order', '2x3 [1,2]=2', f'=INDEX({FN}(2,3),1,2)', 2, None, 'Fill across columns first'),
    ('Fill order', '2x3 [1,3]=3', f'=INDEX({FN}(2,3),1,3)', 3, None, 'End of first row'),
    ('Fill order', '2x3 [2,1]=4', f'=INDEX({FN}(2,3),2,1)', 4, None, 'Start of second row'),
    ('Fill order', '2x3 [2,2]=5', f'=INDEX({FN}(2,3),2,2)', 5, None, 'Middle'),
    ('Fill order', '2x3 [2,3]=6', f'=INDEX({FN}(2,3),2,3)', 6, None, 'Last element'),

    # ── GROUP 10: Interaction with other functions ──────────────────────────
    ('Combos', 'SUM of sequence', f'=SUM({FN}(10))', 55, None, 'SUM(1..10)=55'),
    ('Combos', 'AVERAGE of sequence', f'=AVERAGE({FN}(10))', 5.5, None, 'AVG(1..10)=5.5'),
    ('Combos', 'MAX of sequence', f'=MAX({FN}(5,1,10,3))', 22, None, 'max(10,13,16,19,22)=22'),
    ('Combos', 'MIN of sequence', f'=MIN({FN}(5,1,10,3))', 10, None, 'min=10'),
    ('Combos', 'COUNT of sequence', f'=COUNT({FN}(4,5))', 20, None, '4*5=20 numbers'),

    # ── GROUP 11: Key behavioral questions (confirmed) ──────────────────────
    ('Confirmed', '1048576 rows works', f'=ROWS({FN}(1048576))', 1048576, None, 'Max sheet rows — works'),
    ('Confirmed', '1048577 rows -> #VALUE!', f'={FN}(1048577)', '#VALUE!', 3, 'Exceeds max rows -> #VALUE!'),
    ('Confirmed', '16384 cols works', f'=COLUMNS({FN}(1,16384))', 16384, None, 'Max sheet cols — works'),
    ('Confirmed', '16385 cols -> returns 1', f'={FN}(1,16385)', 1, None, 'Exceeds max cols — returns scalar 1 (bizarre Excel behavior)'),
    ('Confirmed', '1000x1000 = 1M cells', f'=COUNT({FN}(1000,1000))', 1000000, None, '1M cells works'),
    ('Confirmed', 'Spill into occupied cell', None, None, None, 'Manual test: put value in B1, =SEQUENCE(2) in A1 -> #SPILL!'),

    # ── GROUP 12: Dynamic arguments (HyperFormula architectural limitation) ──
    # These rows document Excel behavior for reference. In Excel, cell refs work
    # for dimensions. In HyperFormula, they return #VALUE! because array size
    # must be known at parse time.
    ('Dynamic args', 'rows from cell ref (K3=3)', f'=ROWS({FN}(K3))', 3, None, 'Excel: works. HF: #VALUE! (parse-time size prediction)'),
    ('Dynamic args', 'cols from formula (1+1)', f'=COLUMNS({FN}(3,1+1))', 2, None, 'Excel: works. HF: #VALUE! (parse-time size prediction)'),
]

# ── Write test rows ──────────────────────────────────────────────────────────
row = 3
current_group = None
test_num = 0

for group, desc, formula, expected, error_type, notes in tests:
    # Group header
    if group != current_group:
        current_group = group
        ws.cell(row=row, column=2, value=group).font = group_font
        row += 1

    test_num += 1
    is_manual = formula is None
    is_error_test = error_type is not None

    # Column A: test number
    ws.cell(row=row, column=1, value=test_num)

    # Column B: group name (for filtering)
    ws.cell(row=row, column=2, value=group)

    # Column C: description
    ws.cell(row=row, column=3, value=desc)

    # Column D: formula as text (display only)
    display_formula = formula if formula else '(manual test)'
    ws.cell(row=row, column=4, value=display_formula)

    # Column E: expected value
    if is_manual:
        ws.cell(row=row, column=5, value='INFO')
        ws.cell(row=row, column=5).fill = info_fill
    elif is_error_test:
        # Display the error name, store error_type code in column I (hidden helper)
        ws.cell(row=row, column=5, value=expected)
    else:
        ws.cell(row=row, column=5, value=expected)

    # Column F: actual (live formula)
    if formula:
        ws.cell(row=row, column=6, value=formula)
    else:
        ws.cell(row=row, column=6, value='(manual)')

    # Column G: Pass/Fail
    if is_manual:
        ws.cell(row=row, column=7, value='INFO')
        ws.cell(row=row, column=7).fill = info_fill
    elif is_error_test:
        # Check that actual is an error with the expected ERROR.TYPE code
        actual_ref = f'F{row}'
        ws.cell(row=row, column=7,
                value=f'=IF(ISERROR({actual_ref}),IF(ERROR.TYPE({actual_ref})={error_type},"PASS","FAIL"),"FAIL")')
    else:
        # Compare actual vs expected value
        actual_ref = f'F{row}'
        expected_ref = f'E{row}'
        ws.cell(row=row, column=7, value=f'=IF({actual_ref}={expected_ref},"PASS","FAIL")')

    # Column H: notes
    ws.cell(row=row, column=8, value=notes).font = note_font

    # Pink highlight for manual/info rows
    if is_manual:
        for col in range(1, 9):
            if col not in (5, 7):
                ws.cell(row=row, column=col).fill = info_fill

    # Subtle border on all rows
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

    row += 1

# ── Conditional formatting for Pass/Fail column ─────────────────────────────
pass_fail_range = f'G3:G{row - 1}'
ws.conditional_formatting.add(
    pass_fail_range,
    CellIsRule(operator='equal', formula=['"PASS"'], fill=pass_fill, font=pass_font),
)
ws.conditional_formatting.add(
    pass_fail_range,
    CellIsRule(operator='equal', formula=['"FAIL"'], fill=fail_fill, font=fail_font),
)

# ── Instructions sheet ───────────────────────────────────────────────────────
ws_inst = wb.create_sheet('Instructions')
instructions = [
    'SEQUENCE Validation Workbook — Instructions',
    '',
    '1. Open this file in Excel desktop (2021+ or Microsoft 365)',
    '   - Excel Online may not support _xlfn.SEQUENCE',
    '   - LibreOffice/Google Sheets may behave differently',
    '',
    '2. Go to the "SEQUENCE Validation" sheet',
    '',
    '3. Check the Pass/Fail column (G):',
    '   - PASS (green) = actual matches expected',
    '   - FAIL (red)   = mismatch — investigate!',
    '   - INFO (pink)  = manual/info only, no auto-validation',
    '',
    '4. Error tests use ERROR.TYPE() comparison:',
    '   ERROR.TYPE codes: 2=#DIV/0!, 3=#VALUE!, 6=#NUM!, 7=#N/A, 14=#CALC!',
    '   The formula checks that the actual result is an error with the expected code.',
    '',
    '5. Manual tests (INFO rows):',
    '   - #41, #42: Excel rejects these formulas at parse time (syntax error)',
    '   - #80: Spill test — put value in B1, =SEQUENCE(2) in A1, verify #SPILL!',
    '',
    '6b. Dynamic argument tests (#81-#82):',
    '   These test Excel behavior that HyperFormula handles differently.',
    '   In Excel, cell refs and formulas work for rows/cols dimensions.',
    '   In HyperFormula, they return #VALUE! (array size must be known at parse time).',
    '',
    '6. Setup area (J:K on main sheet) has test fixture values:',
    '   K3=3, K5=0, K6=-1, K13=empty, K14="", K15=#N/A, K16=#DIV/0!',
]
for i, line in enumerate(instructions, start=1):
    ws_inst.cell(row=i, column=1, value=line)
ws_inst.column_dimensions['A'].width = 80

# ── Save ─────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)

auto_count = sum(1 for t in tests if t[2] is not None)
error_count = sum(1 for t in tests if t[4] is not None)
manual_count = sum(1 for t in tests if t[2] is None)

print(f'Written: {OUTPUT_PATH}')
print(f'Total tests: {test_num}')
print(f'Auto-validated: {auto_count} ({auto_count - error_count} value + {error_count} error)')
print(f'Manual/info: {manual_count}')
print()
print('Open in Excel — expect all auto-validated rows to show PASS (green).')
